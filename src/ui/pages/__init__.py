import sys, os
sys.path.insert(0, str(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton,
    QGridLayout, QScrollArea, QLineEdit, QTabWidget, QTextEdit,
    QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit, QSlider,
    QGroupBox, QFormLayout, QListWidget, QListWidgetItem, QMessageBox,
    QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView, QSizePolicy
)
from PySide6.QtCore import Qt, QTimer, QDate, QSize
from PySide6.QtGui import QFont, QPixmap
from datetime import date, timedelta, datetime
from typing import Optional

from src.ui.widgets import (
    Card, StatRow, ProgressRing, MiniButton, AccomplishmentBadge, SectionHeader, STYLE_CARD
)
from src.services import (
    UserService, WeightService, WorkoutService, NutritionService,
    RecoveryService, PRService, AchievementService, OverloadService,
    MuscleHeatmapService, SleepService, MeasurementService
)
from src.ai import AICoach
from src.analytics import ChartBuilder, html_from_figure

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
    HAS_WEBENGINE = True
except ImportError:
    HAS_WEBENGINE = False


class PlotWidget(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setStyleSheet(STYLE_CARD)
        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._view = None
        if HAS_WEBENGINE:
            self._view = QWebEngineView()
            self._view.setStyleSheet("background: transparent; border-radius: 16px;")
            self._layout.addWidget(self._view)

    def set_chart(self, fig):
        if HAS_WEBENGINE and fig is not None:
            html = html_from_figure(fig)
            self._view.setHtml(html)


class DashboardPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()
        self._timer = QTimer()
        self._timer.timeout.connect(self.refresh)
        self._timer.setSingleShot(True)

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        header = QHBoxLayout()
        title = QLabel("Dashboard")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff; letter-spacing: -0.5px;")
        header.addWidget(title)
        header.addStretch()

        self.coach_label = QLabel("")
        self.coach_label.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 rgba(124,58,237,0.12), stop:1 rgba(124,58,237,0.05));
            border: 1px solid rgba(124,58,237,0.2);
            border-radius: 12px;
            padding: 12px 16px;
            font-size: 13px;
            color: #ccc;
        """)
        self.coach_label.setWordWrap(True)
        header.addWidget(self.coach_label, 1)

        refresh_btn = MiniButton("Refresh", "🔄")
        refresh_btn.clicked.connect(self.refresh)
        header.addWidget(refresh_btn)

        layout.addLayout(header)

        cards_grid = QGridLayout()
        cards_grid.setSpacing(12)

        self.card_weight = Card("Current Weight")
        cards_grid.addWidget(self.card_weight, 0, 0)

        self.card_goal = Card("Goal Weight")
        cards_grid.addWidget(self.card_goal, 0, 1)

        self.card_calories = Card("Calories")
        cards_grid.addWidget(self.card_calories, 0, 2)

        self.card_protein = Card("Protein")
        cards_grid.addWidget(self.card_protein, 0, 3)

        self.card_water = Card("Water")
        cards_grid.addWidget(self.card_water, 1, 0)

        self.card_recovery = Card("Recovery")
        cards_grid.addWidget(self.card_recovery, 1, 1)

        self.card_streak = Card("Streak")
        cards_grid.addWidget(self.card_streak, 1, 2)

        self.card_workout = Card("Workout Today")
        cards_grid.addWidget(self.card_workout, 1, 3)

        layout.addLayout(cards_grid)

        mid_row = QHBoxLayout()
        mid_row.setSpacing(16)

        progress_section = QFrame()
        progress_section.setObjectName("card")
        progress_section.setStyleSheet(STYLE_CARD)
        progress_layout = QVBoxLayout(progress_section)
        progress_layout.setContentsMargins(20, 16, 20, 16)
        progress_layout.setSpacing(8)

        progress_title = QLabel("🎯 Goal Progress")
        progress_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #fff;")
        progress_layout.addWidget(progress_title)

        self.progress_ring = ProgressRing(size=160, stroke=10)
        progress_layout.addWidget(self.progress_ring, 0, Qt.AlignCenter)

        self.progress_label = QLabel("63.4 kg → 72.5 kg")
        self.progress_label.setStyleSheet("font-size: 13px; color: #aaa;")
        self.progress_label.setAlignment(Qt.AlignCenter)
        progress_layout.addWidget(self.progress_label)

        self.eta_label = QLabel("")
        self.eta_label.setStyleSheet("font-size: 12px; color: #f59e0b;")
        self.eta_label.setAlignment(Qt.AlignCenter)
        progress_layout.addWidget(self.eta_label)

        mid_row.addWidget(progress_section)

        self.chart_weight = PlotWidget()
        self.chart_weight.setMinimumHeight(280)
        mid_row.addWidget(self.chart_weight, 2)

        self.chart_calories = PlotWidget()
        self.chart_calories.setMinimumHeight(280)
        mid_row.addWidget(self.chart_calories, 2)

        layout.addLayout(mid_row)

        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(16)

        recent_section = QFrame()
        recent_section.setObjectName("card")
        recent_section.setStyleSheet(STYLE_CARD)
        recent_layout = QVBoxLayout(recent_section)
        recent_layout.setContentsMargins(20, 16, 20, 16)

        pr_title = QLabel("🏆 Latest PR")
        pr_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #fff;")
        recent_layout.addWidget(pr_title)

        self.pr_label = QLabel("No PRs yet")
        self.pr_label.setStyleSheet("font-size: 13px; color: #aaa; padding: 8px 0;")
        self.pr_label.setWordWrap(True)
        recent_layout.addWidget(self.pr_label)

        recent_layout.addStretch()
        bottom_row.addWidget(recent_section, 1)

        achievement_section = QFrame()
        achievement_section.setObjectName("card")
        achievement_section.setStyleSheet(STYLE_CARD)
        achievement_layout = QVBoxLayout(achievement_section)
        achievement_layout.setContentsMargins(20, 16, 20, 16)

        ach_title = QLabel("⭐ Recent Achievement")
        ach_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #fff;")
        achievement_layout.addWidget(ach_title)

        self.ach_label = QLabel("")
        self.ach_label.setStyleSheet("font-size: 13px; color: #aaa; padding: 8px 0;")
        self.ach_label.setWordWrap(True)
        achievement_layout.addWidget(self.ach_label)

        achievement_layout.addStretch()
        bottom_row.addWidget(achievement_section, 1)

        layout.addLayout(bottom_row)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        self.refresh()

    def refresh(self):
        user = UserService.get_or_create_user()
        uid = user.id

        latest = WeightService.get_latest(uid)
        current_w = latest.weight_kg if latest else 63.4
        goal_w = user.target_weight_kg
        self.card_weight.set_value(f"{current_w} kg")
        self.card_goal.set_value(f"{goal_w} kg")

        nutrition = NutritionService.get_daily_totals(uid)
        self.card_calories.set_value(f"{nutrition['calories']} / {user.target_calories}")
        protein_pct = int((nutrition['protein'] / user.target_protein_g) * 100) if user.target_protein_g else 0
        self.card_protein.set_value(f"{nutrition['protein']}g / {user.target_protein_g}g")
        self.card_protein.set_subtitle(f"{protein_pct}%")

        water_pct = int((nutrition['water_ml'] / user.target_water_ml) * 100) if user.target_water_ml else 0
        self.card_water.set_value(f"{nutrition['water_ml'] / 1000:.1f}L")
        self.card_water.set_subtitle(f"{water_pct}%")

        recovery = RecoveryService.calculate_recovery_score(uid)
        self.card_recovery.set_value(f"{int(recovery.recovery_score)}%")
        if recovery.recovery_score >= 80:
            self.card_recovery.set_accent("Ready to push!")
        elif recovery.recovery_score >= 60:
            self.card_recovery.set_accent("Good to train")
        else:
            self.card_recovery.set_accent("Need recovery")

        streak = WorkoutService.get_workout_streak(uid)
        self.card_streak.set_value(f"{streak} days")

        today_wo = WorkoutService.get_today_workout(uid)
        if today_wo:
            self.card_workout.set_value(today_wo.name or today_wo.split_type)
            if today_wo.is_completed:
                self.card_workout.set_subtitle("✓ Completed")
            else:
                self.card_workout.set_subtitle("⏳ Pending")
        else:
            self.card_workout.set_value("Rest Day")
            self.card_workout.set_subtitle("No workout scheduled")

        pct = min(100, (current_w / goal_w) * 100) if goal_w > 0 else 0
        self.progress_ring.set_value_animated(pct)
        self.progress_label.setText(f"{current_w} kg → {goal_w} kg")

        eta_days, eta_date = WeightService.get_eta_to_goal(uid, current_w)
        if eta_days > 0:
            self.eta_label.setText(f"ETA: {eta_date}")
        else:
            self.eta_label.setText(eta_date)

        coach = AICoach()
        self.coach_label.setText(f"🤖 {coach.get_daily_recommendation(uid)}")

        latest_pr = PRService.get_latest_pr(uid)
        if latest_pr:
            self.pr_label.setText(f"{latest_pr['exercise_name']}\n{latest_pr['record_type']}: {latest_pr['value']} ({latest_pr['date']})")
        else:
            self.pr_label.setText("Complete a workout to set your first PR!")

        try:
            achievements = AchievementService.check_achievements(uid)
            if achievements:
                self.ach_label.setText(f"🎉 New: {', '.join(achievements)}")
            else:
                earned = AchievementService.get_all_earned(uid)
                if earned:
                    self.ach_label.setText(f"{len(earned)} achievements earned")
                else:
                    self.ach_label.setText("Keep training to earn achievements")
        except Exception:
            self.ach_label.setText("")

        self.chart_weight.set_chart(ChartBuilder.weight_chart(uid))
        self.chart_calories.set_chart(ChartBuilder.calories_chart(uid))


class WorkoutPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("💪 Workout Tracker")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        row1 = QHBoxLayout()
        row1.setSpacing(12)

        self.card_today = Card("Today's Workout")
        row1.addWidget(self.card_today)

        self.card_volume = Card("Weekly Volume")
        row1.addWidget(self.card_volume)

        self.card_workouts = Card("This Week")
        row1.addWidget(self.card_workouts)

        self.card_total = Card("Total Workouts")
        row1.addWidget(self.card_total)

        layout.addLayout(row1)

        self.chart_volume = PlotWidget()
        self.chart_volume.setMinimumHeight(250)
        layout.addWidget(self.chart_volume)

        self.chart_muscle = PlotWidget()
        self.chart_muscle.setMinimumHeight(250)
        layout.addWidget(self.chart_muscle)

        new_btn = MiniButton("New Workout", "➕")
        new_btn.clicked.connect(self._new_workout)
        layout.addWidget(new_btn)

        self.log_section = QFrame()
        self.log_section.setObjectName("card")
        self.log_section.setStyleSheet(STYLE_CARD)
        self.log_layout = QVBoxLayout(self.log_section)
        self.log_layout.setContentsMargins(20, 16, 20, 16)

        log_title = QLabel("Quick Log Exercise")
        log_title.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        self.log_layout.addWidget(log_title)

        form = QFormLayout()
        form.setSpacing(8)

        from src.models import Exercise
        from src.database import get_session
        session = get_session()
        exercises = session.query(Exercise).order_by(Exercise.name).all()
        session.close()

        self.ex_combo = QComboBox()
        self.ex_combo.setStyleSheet("""
            QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QComboBox:hover { border: 1px solid rgba(124,58,237,0.4); }
            QComboBox::drop-down { border: none; }
            QComboBox QAbstractItemView {
                background: #1a1a2e;
                border: 1px solid rgba(255,255,255,0.1);
                selection-background-color: rgba(124,58,237,0.3);
                color: #fff;
            }
        """)
        for ex in exercises:
            self.ex_combo.addItem(ex.name, ex.id)
        form.addRow("Exercise:", self.ex_combo)

        self.weight_spin = QDoubleSpinBox()
        self.weight_spin.setRange(0, 500)
        self.weight_spin.setValue(20)
        self.weight_spin.setSuffix(" kg")
        self.weight_spin.setStyleSheet(self._spin_style())
        form.addRow("Weight:", self.weight_spin)

        self.reps_spin = QSpinBox()
        self.reps_spin.setRange(0, 100)
        self.reps_spin.setValue(10)
        self.reps_spin.setStyleSheet(self._spin_style())
        form.addRow("Reps:", self.reps_spin)

        self.sets_spin = QSpinBox()
        self.sets_spin.setRange(1, 20)
        self.sets_spin.setValue(3)
        self.sets_spin.setStyleSheet(self._spin_style())
        form.addRow("Sets:", self.sets_spin)

        self.rpe_spin = QDoubleSpinBox()
        self.rpe_spin.setRange(0, 10)
        self.rpe_spin.setValue(7)
        self.rpe_spin.setSingleStep(0.5)
        self.rpe_spin.setStyleSheet(self._spin_style())
        form.addRow("RPE:", self.rpe_spin)

        log_btn = MiniButton("Log Exercise", "💾")
        log_btn.clicked.connect(self._log_exercise)
        form.addRow(log_btn)

        self.log_layout.addLayout(form)

        self.ai_suggestion = QLabel("")
        self.ai_suggestion.setStyleSheet("""
            background: rgba(124,58,237,0.08);
            border: 1px solid rgba(124,58,237,0.15);
            border-radius: 8px;
            padding: 10px;
            font-size: 12px;
            color: #bbb;
            margin-top: 8px;
        """)
        self.ai_suggestion.setWordWrap(True)
        self.log_layout.addWidget(self.ai_suggestion)

        self.ex_combo.currentIndexChanged.connect(self._update_ai_suggestion)

        layout.addWidget(self.log_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _spin_style(self):
        return """
            QSpinBox, QDoubleSpinBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QSpinBox:hover, QDoubleSpinBox:hover {
                border: 1px solid rgba(124,58,237,0.4);
            }
        """

    def _update_ai_suggestion(self):
        from src.database import get_session
        from src.models import Exercise
        session = get_session()
        try:
            ex_id = self.ex_combo.currentData()
            if ex_id:
                user = UserService.get_or_create_user()
                coach = AICoach()
                suggestion = coach.get_workout_recommendation(user.id, ex_id)
                self.ai_suggestion.setText(f"🤖 AI Coach: {suggestion}")
        finally:
            session.close()

    def _new_workout(self):
        from src.database import get_session
        from src.models import WorkoutSession
        session = get_session()
        try:
            user = UserService.get_or_create_user()
            today = date.today()
            existing = session.query(WorkoutSession).filter(
                WorkoutSession.user_id == user.id,
                WorkoutSession.date == today
            ).first()
            if existing:
                QMessageBox.information(self, "Info", "Workout already exists for today.")
                return

            wo = WorkoutSession(
                user_id=user.id, date=today,
                split_type=user.workout_split,
                name=f"{user.workout_split} {today.strftime('%a')}"
            )
            session.add(wo)
            session.commit()
            QMessageBox.information(self, "Created", f"Workout created: {wo.name}")
            self.on_show()
        finally:
            session.close()

    def _log_exercise(self):
        from src.database import get_session
        from src.models import WorkoutSession, WorkoutExercise, ExerciseSet, Exercise
        session = get_session()
        try:
            user = UserService.get_or_create_user()
            ex_id = self.ex_combo.currentData()
            weight = self.weight_spin.value()
            reps = self.reps_spin.value()
            num_sets = self.sets_spin.value()
            rpe = self.rpe_spin.value()

            today = date.today()

            wo = session.query(WorkoutSession).filter(
                WorkoutSession.user_id == user.id,
                WorkoutSession.date == today,
                WorkoutSession.is_completed == False
            ).first()

            if not wo:
                split = user.workout_split
                wo = WorkoutSession(
                    user_id=user.id, date=today,
                    split_type=split, name=f"{split} {today.strftime('%a')}"
                )
                session.add(wo)
                session.flush()

            we = WorkoutExercise(
                session_id=wo.id, exercise_id=ex_id,
                order=len(wo.exercises) + 1 if wo.exercises else 1
            )
            session.add(we)
            session.flush()

            for i in range(num_sets):
                es = ExerciseSet(
                    workout_exercise_id=we.id,
                    set_number=i + 1,
                    weight_kg=weight,
                    reps=reps,
                    rpe=rpe,
                    is_warmup=False,
                    is_completed=True,
                )
                session.add(es)

            session.commit()
            QMessageBox.information(self, "Logged", f"{num_sets} sets of {weight}kg × {reps} logged!")

            prs = PRService.check_and_update_prs(user.id, wo.id)
            if prs:
                for pr_type, ex_name, val in prs:
                    QMessageBox.information(self, "🎉 New PR!", f"{pr_type}: {ex_name} = {val}")

            try:
                new_achs = AchievementService.check_achievements(user.id)
                if new_achs:
                    QMessageBox.information(self, "⭐ Achievement!", f"Unlocked: {', '.join(new_achs)}")
            except Exception:
                pass

            self.on_show()
        finally:
            session.close()

    def on_show(self):
        user = UserService.get_or_create_user()
        uid = user.id

        today_wo = WorkoutService.get_today_workout(uid)
        if today_wo:
            self.card_today.set_value(today_wo.name or today_wo.split_type)
            self.card_today.set_subtitle("✓ Completed" if today_wo.is_completed else "⏳ Pending")
        else:
            self.card_today.set_value("Rest Day")
            self.card_today.set_subtitle("No workout scheduled")

        self.card_volume.set_value(f"{int(WorkoutService.get_weekly_volume(uid)):,} kg")
        self.card_workouts.set_value(f"{WorkoutService.get_weekly_workouts_count(uid)} / {user.workout_days_per_week}")
        self.card_total.set_value(f"{WorkoutService.get_total_workouts(uid)}")

        self.chart_volume.set_chart(ChartBuilder.workout_volume_chart(uid))
        self.chart_muscle.set_chart(ChartBuilder.muscle_volume_chart(uid))

        self._update_ai_suggestion()


class ExercisePage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("🏋️ Exercise Database")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        search_row = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search exercises...")
        self.search_input.setStyleSheet("""
            QLineEdit {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px;
                padding: 10px 16px;
                color: #fff;
                font-size: 13px;
            }
            QLineEdit:focus { border: 1px solid rgba(124,58,237,0.5); }
        """)
        self.search_input.textChanged.connect(self._filter)
        search_row.addWidget(self.search_input, 1)

        self.muscle_filter = QComboBox()
        self.muscle_filter.setStyleSheet("""
            QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QComboBox::drop-down { border: none; width: 24px; }
            QComboBox QAbstractItemView {
                background: #1a1a2e;
                border: 1px solid rgba(255,255,255,0.1);
                selection-background-color: rgba(124,58,237,0.3);
                color: #fff;
            }
        """)
        self.muscle_filter.addItem("All Muscles", "")
        from src.models import MuscleGroup
        for m in MuscleGroup:
            self.muscle_filter.addItem(m.value, m.value)
        self.muscle_filter.currentIndexChanged.connect(self._filter)
        search_row.addWidget(self.muscle_filter)

        layout.addLayout(search_row)

        self.ex_list = QVBoxLayout()
        self.ex_list.setSpacing(8)
        layout.addLayout(self.ex_list)
        layout.addStretch()

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _filter(self):
        self._rebuild_list()

    def _rebuild_list(self):
        while self.ex_list.count():
            item = self.ex_list.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        from src.database import get_session
        from src.models import Exercise
        session = get_session()
        try:
            query = session.query(Exercise)
            search = self.search_input.text().strip().lower()
            muscle = self.muscle_filter.currentData()

            if muscle:
                query = query.filter(Exercise.primary_muscle == muscle)

            exercises = query.order_by(Exercise.primary_muscle, Exercise.name).all()

            if search:
                exercises = [e for e in exercises if search in e.name.lower() or search in e.primary_muscle.lower()]

            for ex in exercises:
                card = QFrame()
                card.setObjectName("card")
                card.setStyleSheet(STYLE_CARD)
                card.setCursor(Qt.PointingHandCursor)
                cl = QVBoxLayout(card)
                cl.setContentsMargins(16, 12, 16, 12)
                cl.setSpacing(4)

                name_row = QHBoxLayout()
                name_lbl = QLabel(ex.name)
                name_lbl.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
                name_row.addWidget(name_lbl)
                name_row.addStretch()

                muscle_badge = QLabel(ex.primary_muscle)
                muscle_badge.setStyleSheet("""
                    background: rgba(124,58,237,0.15);
                    color: #7c3aed;
                    border-radius: 6px;
                    padding: 2px 10px;
                    font-size: 11px;
                    font-weight: 600;
                """)
                name_row.addWidget(muscle_badge)

                equip_badge = QLabel(ex.equipment)
                equip_badge.setStyleSheet("""
                    background: rgba(16,185,129,0.15);
                    color: #10b981;
                    border-radius: 6px;
                    padding: 2px 10px;
                    font-size: 11px;
                    font-weight: 600;
                """)
                name_row.addWidget(equip_badge)

                cl.addLayout(name_row)

                if ex.secondary_muscles:
                    sec = QLabel(f"Secondary: {', '.join(ex.secondary_muscles[:3])}")
                    sec.setStyleSheet("font-size: 12px; color: #888;")
                    cl.addWidget(sec)

                if ex.execution:
                    exec_lbl = QLabel(ex.execution[:120] + ("..." if len(ex.execution) > 120 else ""))
                    exec_lbl.setStyleSheet("font-size: 12px; color: #aaa;")
                    exec_lbl.setWordWrap(True)
                    cl.addWidget(exec_lbl)

                details = QLabel(f"Tempo: {ex.recommended_tempo} · Rest: {ex.recommended_rest_seconds}s · Difficulty: {ex.difficulty}")
                details.setStyleSheet("font-size: 11px; color: #666;")
                cl.addWidget(details)

                self.ex_list.addWidget(card)
        finally:
            session.close()

    def on_show(self):
        self._rebuild_list()


class NutritionPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("🥩 Nutrition")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        row1 = QHBoxLayout()
        row1.setSpacing(12)

        self.card_cal = Card("Calories")
        row1.addWidget(self.card_cal)

        self.card_prot = Card("Protein")
        row1.addWidget(self.card_prot)

        self.card_fat = Card("Fat")
        row1.addWidget(self.card_fat)

        self.card_carbs = Card("Carbs")
        row1.addWidget(self.card_carbs)

        self.card_fiber = Card("Fiber")
        row1.addWidget(self.card_fiber)

        self.card_water_nut = Card("Water")
        row1.addWidget(self.card_water_nut)

        layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.setSpacing(12)

        self.card_protein_streak = Card("Protein Streak")
        row2.addWidget(self.card_protein_streak)

        self.card_nutrition_score = Card("Nutrition Score")
        row2.addWidget(self.card_nutrition_score)

        layout.addLayout(row2)

        charts_row = QHBoxLayout()
        charts_row.setSpacing(12)

        self.chart_cal = PlotWidget()
        charts_row.addWidget(self.chart_cal)

        self.chart_prot = PlotWidget()
        charts_row.addWidget(self.chart_prot)

        layout.addLayout(charts_row)

        form_section = QFrame()
        form_section.setObjectName("card")
        form_section.setStyleSheet(STYLE_CARD)
        form_layout = QVBoxLayout(form_section)
        form_layout.setContentsMargins(20, 16, 20, 16)

        form_title = QLabel("Quick Log Meal")
        form_title.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        form_layout.addWidget(form_title)

        f = QFormLayout()
        f.setSpacing(8)

        self.meal_combo = QComboBox()
        self.meal_combo.addItems(["Breakfast", "Lunch", "Dinner", "Snack"])
        self.meal_combo.setStyleSheet(self._combo_style())
        f.addRow("Meal:", self.meal_combo)

        self.food_input = QLineEdit()
        self.food_input.setPlaceholderText("e.g., Chicken breast, rice, broccoli")
        self.food_input.setStyleSheet(self._input_style())
        f.addRow("Food:", self.food_input)

        cal_row = QHBoxLayout()
        self.cal_input = QSpinBox()
        self.cal_input.setRange(0, 5000)
        self.cal_input.setValue(600)
        self.cal_input.setStyleSheet(self._spin_style())
        cal_row.addWidget(QLabel("Calories:"))
        cal_row.addWidget(self.cal_input)

        self.protein_input = QDoubleSpinBox()
        self.protein_input.setRange(0, 200)
        self.protein_input.setValue(30)
        self.protein_input.setSuffix("g")
        self.protein_input.setStyleSheet(self._spin_style())
        cal_row.addWidget(QLabel("Protein:"))
        cal_row.addWidget(self.protein_input)

        self.fat_input = QDoubleSpinBox()
        self.fat_input.setRange(0, 200)
        self.fat_input.setValue(15)
        self.fat_input.setSuffix("g")
        self.fat_input.setStyleSheet(self._spin_style())
        cal_row.addWidget(QLabel("Fat:"))
        cal_row.addWidget(self.fat_input)

        self.carbs_input = QDoubleSpinBox()
        self.carbs_input.setRange(0, 500)
        self.carbs_input.setValue(60)
        self.carbs_input.setSuffix("g")
        self.carbs_input.setStyleSheet(self._spin_style())
        cal_row.addWidget(QLabel("Carbs:"))
        cal_row.addWidget(self.carbs_input)

        f.addRow(cal_row)

        log_btn = MiniButton("Log Meal", "💾")
        log_btn.clicked.connect(self._log_meal)
        f.addRow(log_btn)

        water_row = QHBoxLayout()
        self.water_input = QSpinBox()
        self.water_input.setRange(0, 5000)
        self.water_input.setValue(250)
        self.water_input.setSuffix(" ml")
        self.water_input.setStyleSheet(self._spin_style())
        water_row.addWidget(QLabel("Water:"))
        water_row.addWidget(self.water_input)

        water_btn = MiniButton("Log Water", "💧")
        water_btn.clicked.connect(self._log_water)
        water_row.addWidget(water_btn)

        f.addRow(water_row)

        self.ai_nutrition = QLabel("")
        self.ai_nutrition.setStyleSheet("""
            background: rgba(124,58,237,0.08);
            border: 1px solid rgba(124,58,237,0.15);
            border-radius: 8px;
            padding: 10px;
            font-size: 12px;
            color: #bbb;
            margin-top: 8px;
        """)
        self.ai_nutrition.setWordWrap(True)
        form_layout.addWidget(self.ai_nutrition)

        form_layout.addLayout(f)
        layout.addWidget(form_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _combo_style(self):
        return """
            QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QComboBox::drop-down { border: none; width: 24px; }
            QComboBox QAbstractItemView {
                background: #1a1a2e;
                border: 1px solid rgba(255,255,255,0.1);
                selection-background-color: rgba(124,58,237,0.3);
                color: #fff;
            }
        """

    def _input_style(self):
        return """
            QLineEdit {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QLineEdit:focus { border: 1px solid rgba(124,58,237,0.5); }
        """

    def _spin_style(self):
        return """
            QSpinBox, QDoubleSpinBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 6px;
                padding: 6px 10px;
                color: #fff;
                font-size: 12px;
            }
        """

    def _log_meal(self):
        user = UserService.get_or_create_user()
        NutritionService.log_meal(
            user.id,
            self.meal_combo.currentText().lower(),
            self.cal_input.value(),
            self.protein_input.value(),
            self.fat_input.value(),
            self.carbs_input.value(),
            food_name=self.food_input.text()
        )
        self.food_input.clear()
        self.on_show()

    def _log_water(self):
        user = UserService.get_or_create_user()
        NutritionService.log_water(user.id, self.water_input.value())
        self.on_show()

    def on_show(self):
        user = UserService.get_or_create_user()
        uid = user.id

        nutrition = NutritionService.get_daily_totals(uid)
        cal_pct = int((nutrition['calories'] / user.target_calories) * 100) if user.target_calories else 0
        prot_pct = int((nutrition['protein'] / user.target_protein_g) * 100) if user.target_protein_g else 0
        fat_pct = int((nutrition['fat'] / user.target_fat_g) * 100) if user.target_fat_g else 0
        carbs_pct = int((nutrition['carbs'] / user.target_carbs_g) * 100) if user.target_carbs_g else 0
        fiber_pct = int((nutrition['fiber'] / user.target_fiber_g) * 100) if user.target_fiber_g else 0
        water_pct = int((nutrition['water_ml'] / user.target_water_ml) * 100) if user.target_water_ml else 0

        self.card_cal.set_value(f"{nutrition['calories']}")
        self.card_cal.set_subtitle(f"Target: {user.target_calories} ({cal_pct}%)")

        self.card_prot.set_value(f"{nutrition['protein']}g")
        self.card_prot.set_subtitle(f"Target: {user.target_protein_g}g ({prot_pct}%)")

        self.card_fat.set_value(f"{nutrition['fat']}g")
        self.card_fat.set_subtitle(f"Target: {user.target_fat_g}g ({fat_pct}%)")

        self.card_carbs.set_value(f"{nutrition['carbs']}g")
        self.card_carbs.set_subtitle(f"Target: {user.target_carbs_g}g ({carbs_pct}%)")

        self.card_fiber.set_value(f"{nutrition['fiber']}g")
        self.card_fiber.set_subtitle(f"Target: {user.target_fiber_g}g ({fiber_pct}%)")

        self.card_water_nut.set_value(f"{nutrition['water_ml'] / 1000:.1f}L")
        self.card_water_nut.set_subtitle(f"Target: {user.target_water_ml / 1000:.1f}L ({water_pct}%)")

        protein_streak = NutritionService.get_daily_protein_streak(uid)
        self.card_protein_streak.set_value(f"{protein_streak} days")

        nutrition_score = int((prot_pct * 0.4 + cal_pct * 0.3 + water_pct * 0.3))
        self.card_nutrition_score.set_value(f"{nutrition_score}%")

        coach = AICoach()
        self.ai_nutrition.setText(f"🤖 {coach.get_nutrition_recommendation(uid)}")

        self.chart_cal.set_chart(ChartBuilder.calories_chart(uid))
        self.chart_prot.set_chart(ChartBuilder.protein_chart(uid))


class WeightPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("⚖️ Weight Tracker")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        row1 = QHBoxLayout()
        row1.setSpacing(12)

        self.card_current = Card("Current")
        row1.addWidget(self.card_current)

        self.card_weekly = Card("Weekly Avg")
        row1.addWidget(self.card_weekly)

        self.card_monthly = Card("Monthly Avg")
        row1.addWidget(self.card_monthly)

        self.card_eta = Card("ETA to Goal")
        row1.addWidget(self.card_eta)

        layout.addLayout(row1)

        self.chart_weight_w = PlotWidget()
        self.chart_weight_w.setMinimumHeight(300)
        layout.addWidget(self.chart_weight_w)

        log_section = QFrame()
        log_section.setObjectName("card")
        log_section.setStyleSheet(STYLE_CARD)
        log_form = QHBoxLayout(log_section)
        log_form.setContentsMargins(16, 12, 16, 12)

        log_form.addWidget(QLabel("Log Weight:"))
        self.weight_input = QDoubleSpinBox()
        self.weight_input.setRange(30, 200)
        self.weight_input.setValue(63.4)
        self.weight_input.setDecimals(1)
        self.weight_input.setSuffix(" kg")
        self.weight_input.setStyleSheet("""
            QDoubleSpinBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 14px;
            }
        """)
        log_form.addWidget(self.weight_input)

        log_btn = MiniButton("Log", "💾")
        log_btn.clicked.connect(self._log_weight)
        log_form.addWidget(log_btn)

        log_form.addStretch()
        layout.addWidget(log_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _log_weight(self):
        user = UserService.get_or_create_user()
        WeightService.log_weight(user.id, self.weight_input.value())
        self.on_show()

    def on_show(self):
        user = UserService.get_or_create_user()
        uid = user.id

        latest = WeightService.get_latest(uid)
        current = latest.weight_kg if latest else 63.4

        self.card_current.set_value(f"{current} kg")
        self.card_current.set_subtitle(f"Goal: {user.target_weight_kg} kg")

        weekly = WeightService.get_weekly_average(uid)
        self.card_weekly.set_value(f"{weekly} kg" if weekly else "No data")

        monthly = WeightService.get_monthly_average(uid)
        self.card_monthly.set_value(f"{monthly} kg" if monthly else "No data")

        eta_days, eta_date = WeightService.get_eta_to_goal(uid, current)
        if eta_days > 0:
            self.card_eta.set_value(f"{eta_days} days")
            self.card_eta.set_subtitle(eta_date)
        else:
            self.card_eta.set_value(eta_date)

        self.chart_weight_w.set_chart(ChartBuilder.weight_chart(uid))


class PhysiquePage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("📸 Physique Tracker")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        measure_section = QFrame()
        measure_section.setObjectName("card")
        measure_section.setStyleSheet(STYLE_CARD)
        measure_layout = QVBoxLayout(measure_section)
        measure_layout.setContentsMargins(20, 16, 20, 16)

        mt = QLabel("📏 Measurements")
        mt.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        measure_layout.addWidget(mt)

        self.measure_cards = {}
        measures = [
            ("chest_cm", "Chest"), ("waist_cm", "Waist"), ("shoulders_cm", "Shoulders"),
            ("left_arm_cm", "Left Arm"), ("right_arm_cm", "Right Arm"),
            ("left_thigh_cm", "Left Thigh"), ("right_thigh_cm", "Right Thigh"),
            ("left_calf_cm", "Left Calf"), ("right_calf_cm", "Right Calf"),
            ("body_fat_pct", "Body Fat %"),
        ]

        grid = QGridLayout()
        grid.setSpacing(8)
        for i, (key, label) in enumerate(measures):
            card = Card(label)
            self.measure_cards[key] = card
            grid.addWidget(card, i // 3, i % 3)

        measure_layout.addLayout(grid)

        form = QHBoxLayout()
        self.measure_inputs = {}
        for key, label in measures[:5]:
            inp = QDoubleSpinBox()
            inp.setRange(0, 200)
            inp.setDecimals(1)
            inp.setStyleSheet("""
                QDoubleSpinBox {
                    background: rgba(255,255,255,0.06);
                    border: 1px solid rgba(255,255,255,0.1);
                    border-radius: 6px;
                    padding: 4px 8px;
                    color: #fff;
                    font-size: 11px;
                    max-width: 100px;
                }
            """)
            self.measure_inputs[key] = inp
            form.addWidget(QLabel(label.split()[-1] + ":"))
            form.addWidget(inp)

        measure_layout.addLayout(form)

        log_btn = MiniButton("Log Measurements", "💾")
        log_btn.clicked.connect(self._log_measurements)
        measure_layout.addWidget(log_btn)

        layout.addWidget(measure_section)

        photo_section = QFrame()
        photo_section.setObjectName("card")
        photo_section.setStyleSheet(STYLE_CARD)
        photo_layout = QVBoxLayout(photo_section)
        photo_layout.setContentsMargins(20, 16, 20, 16)

        pt = QLabel("📸 Physique Photos")
        pt.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        photo_layout.addWidget(pt)

        photo_info = QLabel("Take front, side, and back photos monthly to track visual progress.")
        photo_info.setStyleSheet("font-size: 13px; color: #888;")
        photo_info.setWordWrap(True)
        photo_layout.addWidget(photo_info)

        photo_btn = MiniButton("Import Photo", "📁")
        photo_btn.clicked.connect(self._import_photo)
        photo_layout.addWidget(photo_btn)

        layout.addWidget(photo_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _log_measurements(self):
        user = UserService.get_or_create_user()
        kwargs = {k: v.value() for k, v in self.measure_inputs.items()}
        MeasurementService.log_measurements(user.id, **kwargs)
        self.on_show()

    def _import_photo(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(self, "Select Photo", "", "Images (*.png *.jpg *.jpeg)")
        if path:
            from src.models import PhysiquePhoto
            from src.database import get_session
            import shutil
            dest = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), "images", os.path.basename(path))
            shutil.copy2(path, dest)
            user = UserService.get_or_create_user()
            session = get_session()
            try:
                pp = PhysiquePhoto(user_id=user.id, file_path=dest, angle="front")
                session.add(pp)
                session.commit()
            finally:
                session.close()

    def on_show(self):
        user = UserService.get_or_create_user()
        latest = MeasurementService.get_latest(user.id)
        if latest:
            for key, card in self.measure_cards.items():
                val = getattr(latest, key, 0)
                suffix = "%" if key == "body_fat_pct" else " cm"
                card.set_value(f"{val}{suffix}" if val else "—")


class PRPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("🏆 Personal Records")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        self.pr_list = QVBoxLayout()
        self.pr_list.setSpacing(8)
        layout.addLayout(self.pr_list)
        layout.addStretch()

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        while self.pr_list.count():
            item = self.pr_list.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        user = UserService.get_or_create_user()
        prs = PRService.get_all_prs(user.id)

        if not prs:
            lbl = QLabel("No personal records yet. Complete workouts to set PRs!")
            lbl.setStyleSheet("font-size: 14px; color: #666; padding: 20px;")
            self.pr_list.addWidget(lbl)
            return

        for pr in prs:
            card = QFrame()
            card.setObjectName("card")
            card.setStyleSheet(STYLE_CARD)
            cl = QHBoxLayout(card)
            cl.setContentsMargins(16, 12, 16, 12)

            icon_map = {"weight_1rm": "💪", "best_set": "📊", "best_reps": "🔁"}
            icon_lbl = QLabel(icon_map.get(pr["record_type"], "🏆"))
            icon_lbl.setStyleSheet("font-size: 24px;")
            cl.addWidget(icon_lbl)

            info = QVBoxLayout()
            name_lbl = QLabel(pr["exercise_name"])
            name_lbl.setStyleSheet("font-size: 15px; font-weight: 700; color: #fff;")
            info.addWidget(name_lbl)

            type_lbl = QLabel(f"{pr['record_type']} · {pr['value']} · {pr['date']}")
            type_lbl.setStyleSheet("font-size: 12px; color: #888;")
            info.addWidget(type_lbl)
            cl.addLayout(info, 1)

            val_lbl = QLabel(str(pr["value"]))
            val_lbl.setStyleSheet("font-size: 20px; font-weight: 800; color: #f59e0b;")
            cl.addWidget(val_lbl)

            self.pr_list.addWidget(card)


class MuscleHeatmapPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("🔥 Muscle Heatmap")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        info = QLabel("Volume distribution across muscle groups. Shows undertrained, optimal, and overtrained muscles.")
        info.setStyleSheet("font-size: 13px; color: #888;")
        info.setWordWrap(True)
        layout.addWidget(info)

        self.heatmap_grid = QGridLayout()
        self.heatmap_grid.setSpacing(10)
        layout.addLayout(self.heatmap_grid)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        while self.heatmap_grid.count():
            item = self.heatmap_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        user = UserService.get_or_create_user()
        data = MuscleHeatmapService.get_heatmap_data(user.id)

        color_map = {
            "untrained": "#374151",
            "undertrained": "#ef4444",
            "optimal": "#10b981",
            "overtrained": "#f59e0b",
        }

        for i, (muscle, d) in enumerate(sorted(data.items(), key=lambda x: x[1]["percentage"], reverse=True)):
            card = QFrame()
            card.setObjectName("card")
            card.setStyleSheet(STYLE_CARD)
            card.setMinimumSize(160, 100)

            cl = QVBoxLayout(card)
            cl.setContentsMargins(14, 10, 14, 10)
            cl.setSpacing(4)

            name_lbl = QLabel(muscle)
            name_lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #fff;")
            cl.addWidget(name_lbl)

            color = color_map.get(d["status"], "#666")
            vol_lbl = QLabel(f"{int(d['volume']):,} / {int(d['target']):,} kg")
            vol_lbl.setStyleSheet(f"font-size: 12px; color: {color};")
            cl.addWidget(vol_lbl)

            pct = d["percentage"]
            bar = QFrame()
            bar.setStyleSheet(f"""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 {color}44);
                border-radius: 3px;
                min-height: 6px;
                max-height: 6px;
            """)
            bar.setFixedWidth(int(pct * 1.4) if pct < 100 else 140)
            cl.addWidget(bar)

            status_lbl = QLabel(d["status"].title())
            status_lbl.setStyleSheet(f"font-size: 10px; color: {color}; font-weight: 600; text-transform: uppercase;")
            cl.addWidget(status_lbl)

            self.heatmap_grid.addWidget(card, i // 4, i % 4)


class RecoveryPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("🔄 Recovery")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        row1 = QHBoxLayout()
        row1.setSpacing(12)

        self.card_rec_score = Card("Recovery Score")
        row1.addWidget(self.card_rec_score)

        self.card_fatigue = Card("Fatigue")
        row1.addWidget(self.card_fatigue)

        self.card_sleep_rec = Card("Sleep Score")
        row1.addWidget(self.card_sleep_rec)

        self.card_nutrition_rec = Card("Nutrition Score")
        row1.addWidget(self.card_nutrition_rec)

        self.card_volume_rec = Card("Volume Score")
        row1.addWidget(self.card_volume_rec)

        layout.addLayout(row1)

        self.recovery_rec = QLabel("")
        self.recovery_rec.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 rgba(124,58,237,0.12), stop:1 rgba(124,58,237,0.05));
            border: 1px solid rgba(124,58,237,0.2);
            border-radius: 12px;
            padding: 14px 18px;
            font-size: 14px;
            color: #ccc;
        """)
        self.recovery_rec.setWordWrap(True)
        layout.addWidget(self.recovery_rec)

        self.chart_rec = PlotWidget()
        self.chart_rec.setMinimumHeight(250)
        layout.addWidget(self.chart_rec)

        sleep_section = QFrame()
        sleep_section.setObjectName("card")
        sleep_section.setStyleSheet(STYLE_CARD)
        sleep_layout = QVBoxLayout(sleep_section)
        sleep_layout.setContentsMargins(20, 16, 20, 16)

        st = QLabel("😴 Log Sleep")
        st.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        sleep_layout.addWidget(st)

        sf = QHBoxLayout()
        sf.addWidget(QLabel("Hours:"))
        self.sleep_hours = QDoubleSpinBox()
        self.sleep_hours.setRange(0, 24)
        self.sleep_hours.setValue(8)
        self.sleep_hours.setStyleSheet(self._spin_style())
        sf.addWidget(self.sleep_hours)

        sf.addWidget(QLabel("Quality (1-5):"))
        self.sleep_quality = QSpinBox()
        self.sleep_quality.setRange(1, 5)
        self.sleep_quality.setValue(3)
        self.sleep_quality.setStyleSheet(self._spin_style())
        sf.addWidget(self.sleep_quality)

        sleep_btn = MiniButton("Log Sleep", "💾")
        sleep_btn.clicked.connect(self._log_sleep)
        sf.addWidget(sleep_btn)

        sf.addStretch()
        sleep_layout.addLayout(sf)
        layout.addWidget(sleep_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _spin_style(self):
        return """
            QDoubleSpinBox, QSpinBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 6px;
                padding: 6px 10px;
                color: #fff;
                font-size: 12px;
            }
        """

    def _log_sleep(self):
        user = UserService.get_or_create_user()
        SleepService.log_sleep(user.id, self.sleep_hours.value(), self.sleep_quality.value())
        self.on_show()

    def on_show(self):
        user = UserService.get_or_create_user()
        recovery = RecoveryService.calculate_recovery_score(user.id)

        self.card_rec_score.set_value(f"{int(recovery.recovery_score)}%")
        if recovery.recovery_score >= 80:
            self.card_rec_score.set_accent("Ready!")
        elif recovery.recovery_score >= 60:
            self.card_rec_score.set_accent("Good")
        else:
            self.card_rec_score.set_accent("Rest needed")

        self.card_fatigue.set_value(f"{recovery.fatigue_score}/10")
        self.card_sleep_rec.set_value(f"{int(recovery.sleep_score)}%")
        self.card_nutrition_rec.set_value(f"{int(recovery.nutrition_score)}%")
        self.card_volume_rec.set_value(f"{int(recovery.volume_score)}%")

        self.recovery_rec.setText(f"🔄 {recovery.recommendation}")

        self.chart_rec.set_chart(ChartBuilder.recovery_chart(user.id))


class CalendarPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("📅 Workout Calendar")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        self.card_streak_cal = Card("Current Streak")
        layout.addWidget(self.card_streak_cal)

        self.cal_grid = QGridLayout()
        self.cal_grid.setSpacing(6)
        layout.addLayout(self.cal_grid)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        while self.cal_grid.count():
            item = self.cal_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        user = UserService.get_or_create_user()
        streak = WorkoutService.get_workout_streak(user.id)
        self.card_streak_cal.set_value(f"{streak} days")

        from src.database import get_session
        from src.models import WorkoutSession
        session = get_session()
        try:
            today = date.today()
            month_start = today.replace(day=1)
            if month_start.month == 12:
                month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = today.replace(month=month_start.month + 1, day=1) - timedelta(days=1)

            workouts = session.query(WorkoutSession).filter(
                WorkoutSession.user_id == user.id,
                WorkoutSession.date >= month_start,
                WorkoutSession.date <= month_end,
                WorkoutSession.is_completed == True
            ).all()

            workout_dates = {w.date for w in workouts}

            import calendar as cal_mod
            cal = cal_mod.monthcalendar(today.year, today.month)
            days_header = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            for i, d in enumerate(days_header):
                lbl = QLabel(d)
                lbl.setAlignment(Qt.AlignCenter)
                lbl.setStyleSheet("font-size: 11px; color: #666; font-weight: 600; padding: 4px;")
                self.cal_grid.addWidget(lbl, 0, i)

            for r, week in enumerate(cal):
                for c, day in enumerate(week):
                    if day == 0:
                        empty = QLabel("")
                        empty.setFixedSize(60, 50)
                        self.cal_grid.addWidget(empty, r + 1, c)
                    else:
                        d = date(today.year, today.month, day)
                        is_today = d == today
                        has_wo = d in workout_dates

                        cell = QFrame()
                        cell.setFixedSize(60, 50)

                        bg = "#7c3aed" if has_wo else ("rgba(124,58,237,0.15)" if is_today else "transparent")

                        cell.setStyleSheet(f"""
                            background: {bg};
                            border-radius: 8px;
                            border: {'1px solid #7c3aed' if is_today else '1px solid transparent'};
                        """)

                        cl = QVBoxLayout(cell)
                        cl.setContentsMargins(2, 2, 2, 2)
                        cl.setAlignment(Qt.AlignCenter)

                        day_lbl = QLabel(str(day))
                        day_lbl.setAlignment(Qt.AlignCenter)
                        day_lbl.setStyleSheet(f"""
                            font-size: 13px;
                            font-weight: {'700' if is_today or has_wo else '400'};
                            color: {'#fff' if has_wo else '#999'};
                        """)
                        cl.addWidget(day_lbl)

                        if has_wo:
                            dot = QLabel("●")
                            dot.setAlignment(Qt.AlignCenter)
                            dot.setStyleSheet("font-size: 6px; color: #7c3aed;")
                            cl.addWidget(dot)

                        self.cal_grid.addWidget(cell, r + 1, c)
        finally:
            session.close()


class AnalyticsPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("📈 Analytics")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        self.charts = {}

        chart_configs = [
            ("weight", "Weight Progress", ChartBuilder.weight_chart),
            ("calories", "Calories", ChartBuilder.calories_chart),
            ("protein", "Protein", ChartBuilder.protein_chart),
            ("volume", "Workout Volume", ChartBuilder.workout_volume_chart),
            ("muscle", "Muscle Volume", ChartBuilder.muscle_volume_chart),
            ("recovery", "Recovery", ChartBuilder.recovery_chart),
            ("frequency", "Workout Frequency", ChartBuilder.workout_frequency_chart),
            ("sleep", "Sleep", ChartBuilder.sleep_chart),
            ("water", "Water", ChartBuilder.water_chart),
        ]

        for key, label, builder in chart_configs:
            section = QFrame()
            section.setObjectName("card")
            section.setStyleSheet(STYLE_CARD)
            sl = QVBoxLayout(section)
            sl.setContentsMargins(16, 12, 16, 12)

            lbl = QLabel(label)
            lbl.setStyleSheet("font-size: 14px; font-weight: 700; color: #fff;")
            sl.addWidget(lbl)

            plot = PlotWidget()
            plot.setMinimumHeight(250)
            sl.addWidget(plot)

            self.charts[key] = plot
            layout.addWidget(section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        user = UserService.get_or_create_user()
        uid = user.id

        chart_map = {
            "weight": ChartBuilder.weight_chart,
            "calories": ChartBuilder.calories_chart,
            "protein": ChartBuilder.protein_chart,
            "volume": ChartBuilder.workout_volume_chart,
            "muscle": ChartBuilder.muscle_volume_chart,
            "recovery": ChartBuilder.recovery_chart,
            "frequency": ChartBuilder.workout_frequency_chart,
            "sleep": ChartBuilder.sleep_chart,
            "water": ChartBuilder.water_chart,
        }

        for key, builder in chart_map.items():
            if key in self.charts:
                try:
                    self.charts[key].set_chart(builder(uid))
                except Exception:
                    pass


class AchievementsPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("⭐ Achievements")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        self.ach_list = QVBoxLayout()
        self.ach_list.setSpacing(8)
        layout.addLayout(self.ach_list)
        layout.addStretch()

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def on_show(self):
        while self.ach_list.count():
            item = self.ach_list.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        user = UserService.get_or_create_user()
        try:
            AchievementService.check_achievements(user.id)
        except Exception:
            pass
        earned = AchievementService.get_all_earned(user.id)

        if not earned:
            lbl = QLabel("No achievements yet. Keep training!")
            lbl.setStyleSheet("font-size: 14px; color: #666; padding: 20px;")
            self.ach_list.addWidget(lbl)
            return

        for ach in earned:
            card = QFrame()
            card.setObjectName("card")
            card.setStyleSheet(STYLE_CARD)

            cl = QHBoxLayout(card)
            cl.setContentsMargins(16, 12, 16, 12)

            icon = QLabel(ach["icon"])
            icon.setStyleSheet("font-size: 28px;")
            cl.addWidget(icon)

            info = QVBoxLayout()
            name = QLabel(ach["name"])
            name.setStyleSheet("font-size: 15px; font-weight: 700; color: #fff;")
            info.addWidget(name)

            desc = QLabel(ach["description"])
            desc.setStyleSheet("font-size: 12px; color: #888;")
            info.addWidget(desc)
            cl.addLayout(info, 1)

            date_str = ach["unlocked_at"].strftime("%b %d, %Y") if hasattr(ach["unlocked_at"], "strftime") else str(ach["unlocked_at"])
            date_lbl = QLabel(f"Earned: {date_str}")
            date_lbl.setStyleSheet("font-size: 11px; color: #555;")
            cl.addWidget(date_lbl)

            self.ach_list.addWidget(card)


class SettingsPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background: transparent; border: none;")
        container = QWidget()
        container.setStyleSheet("background: transparent;")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        title = QLabel("⚙️ Settings")
        title.setStyleSheet("font-size: 28px; font-weight: 800; color: #fff;")
        layout.addWidget(title)

        user_section = QFrame()
        user_section.setObjectName("card")
        user_section.setStyleSheet(STYLE_CARD)
        user_layout = QVBoxLayout(user_section)
        user_layout.setContentsMargins(20, 16, 20, 16)
        user_layout.setSpacing(10)

        ut = QLabel("👤 User Profile")
        ut.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        user_layout.addWidget(ut)

        f = QFormLayout()
        f.setSpacing(8)

        self.name_input = QLineEdit()
        self.name_input.setStyleSheet(self._input_style())
        f.addRow("Name:", self.name_input)

        self.age_input = QSpinBox()
        self.age_input.setRange(10, 100)
        self.age_input.setStyleSheet(self._spin_style())
        f.addRow("Age:", self.age_input)

        self.height_input = QDoubleSpinBox()
        self.height_input.setRange(100, 250)
        self.height_input.setSuffix(" cm")
        self.height_input.setStyleSheet(self._spin_style())
        f.addRow("Height:", self.height_input)

        self.goal_input = QComboBox()
        self.goal_input.addItems(["Lean Bulk", "Bulk", "Cut", "Maintenance", "Recomp"])
        self.goal_input.setStyleSheet(self._combo_style())
        f.addRow("Goal:", self.goal_input)

        self.split_input = QComboBox()
        self.split_input.addItems(["PPL-UL", "Push-Pull-Legs", "Upper-Lower", "Full Body", "Bro Split"])
        self.split_input.setStyleSheet(self._combo_style())
        f.addRow("Split:", self.split_input)

        user_layout.addLayout(f)

        target_section = QGroupBox("🏋️ Targets")
        target_section.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: 600;
                color: #fff;
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 20px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
            }
        """)
        tf = QFormLayout(target_section)
        tf.setSpacing(8)

        self.target_weight_input = QDoubleSpinBox()
        self.target_weight_input.setRange(40, 200)
        self.target_weight_input.setSuffix(" kg")
        self.target_weight_input.setStyleSheet(self._spin_style())
        tf.addRow("Target Weight:", self.target_weight_input)

        self.target_cal_input = QSpinBox()
        self.target_cal_input.setRange(1000, 10000)
        self.target_cal_input.setSuffix(" kcal")
        self.target_cal_input.setStyleSheet(self._spin_style())
        tf.addRow("Target Calories:", self.target_cal_input)

        self.target_protein_input = QSpinBox()
        self.target_protein_input.setRange(50, 500)
        self.target_protein_input.setSuffix(" g")
        self.target_protein_input.setStyleSheet(self._spin_style())
        tf.addRow("Target Protein:", self.target_protein_input)

        self.target_water_input = QSpinBox()
        self.target_water_input.setRange(500, 10000)
        self.target_water_input.setSuffix(" ml")
        self.target_water_input.setStyleSheet(self._spin_style())
        tf.addRow("Target Water:", self.target_water_input)

        self.days_input = QSpinBox()
        self.days_input.setRange(1, 7)
        self.days_input.setStyleSheet(self._spin_style())
        tf.addRow("Workout Days/Week:", self.days_input)

        user_layout.addWidget(target_section)

        save_btn = MiniButton("Save Settings", "💾")
        save_btn.clicked.connect(self._save)
        user_layout.addWidget(save_btn)

        layout.addWidget(user_section)

        export_section = QFrame()
        export_section.setObjectName("card")
        export_section.setStyleSheet(STYLE_CARD)
        export_layout = QVBoxLayout(export_section)
        export_layout.setContentsMargins(20, 16, 20, 16)

        et = QLabel("📤 Export Data")
        et.setStyleSheet("font-size: 16px; font-weight: 700; color: #fff;")
        export_layout.addWidget(et)

        export_info = QLabel("Export your data for backup or analysis.")
        export_info.setStyleSheet("font-size: 13px; color: #888;")
        export_layout.addWidget(export_info)

        export_row = QHBoxLayout()
        for fmt, icon in [("CSV", "📄"), ("Excel", "📊"), ("JSON", "📋")]:
            btn = MiniButton(f"Export {fmt}", icon)
            btn.clicked.connect(lambda checked, f=fmt.lower(): self._export(f))
            export_row.addWidget(btn)
        export_row.addStretch()
        export_layout.addLayout(export_row)

        layout.addWidget(export_section)

        scroll.setWidget(container)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _input_style(self):
        return """
            QLineEdit {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QLineEdit:focus { border: 1px solid rgba(124,58,237,0.5); }
        """

    def _combo_style(self):
        return """
            QComboBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 8px;
                padding: 8px 12px;
                color: #fff;
                font-size: 13px;
            }
            QComboBox::drop-down { border: none; width: 24px; }
            QComboBox QAbstractItemView {
                background: #1a1a2e;
                border: 1px solid rgba(255,255,255,0.1);
                selection-background-color: rgba(124,58,237,0.3);
                color: #fff;
            }
        """

    def _spin_style(self):
        return """
            QSpinBox, QDoubleSpinBox {
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 6px;
                padding: 6px 10px;
                color: #fff;
                font-size: 12px;
            }
        """

    def _save(self):
        user = UserService.get_or_create_user()
        UserService.update_user(
            user.id,
            name=self.name_input.text(),
            age=self.age_input.value(),
            height_cm=self.height_input.value(),
            target_weight_kg=self.target_weight_input.value(),
            goal=self.goal_input.currentText(),
            workout_split=self.split_input.currentText(),
            target_calories=self.target_cal_input.value(),
            target_protein_g=self.target_protein_input.value(),
            target_water_ml=self.target_water_input.value(),
            workout_days_per_week=self.days_input.value(),
        )
        QMessageBox.information(self, "Settings", "Settings saved successfully!")

    def _export(self, fmt: str):
        from src.database import get_session
        import csv, json
        from pathlib import Path

        export_dir = Path(__file__).parent.parent.parent.parent / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        user = UserService.get_or_create_user()
        session = get_session()

        try:
            if fmt == "csv":
                path = export_dir / f"gymos_export_{date.today()}.csv"
                with open(path, "w", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["Date", "Weight", "Calories", "Protein", "Fat", "Carbs", "Water", "Workout", "Volume"])

                    from src.models import WeightEntry, NutritionLog, WaterLog, WorkoutSession, WorkoutExercise, ExerciseSet
                    weights = {w.date: w.weight_kg for w in session.query(WeightEntry).filter(WeightEntry.user_id == user.id).all()}
                    cals_data = {}
                    for n in session.query(NutritionLog).filter(NutritionLog.user_id == user.id).all():
                        d = n.date
                        if d not in cals_data:
                            cals_data[d] = {"cal": 0, "pro": 0, "fat": 0, "carbs": 0}
                        cals_data[d]["cal"] += n.calories
                        cals_data[d]["pro"] += n.protein_g
                        cals_data[d]["fat"] += n.fat_g
                        cals_data[d]["carbs"] += n.carbs_g
                    waters = {}
                    for wl in session.query(WaterLog).filter(WaterLog.user_id == user.id).all():
                        waters[wl.date] = waters.get(wl.date, 0) + wl.amount_ml
                    workouts = {}
                    for ws in session.query(WorkoutSession).filter(WorkoutSession.user_id == user.id, WorkoutSession.is_completed == True).all():
                        vol = 0
                        for we in ws.exercises:
                            for es in we.sets:
                                if not es.is_warmup:
                                    vol += es.weight_kg * es.reps
                        workouts[ws.date] = vol

                    all_dates = set(list(weights.keys()) + list(cals_data.keys()) + list(waters.keys()) + list(workouts.keys()))
                    for d in sorted(all_dates):
                        w = weights.get(d, "")
                        c = cals_data.get(d, {})
                        wt = waters.get(d, "")
                        vol = workouts.get(d, "")
                        w.writerow([d, w, c.get("cal", ""), c.get("pro", ""), c.get("fat", ""), c.get("carbs", ""), wt, "Yes" if d in workouts else "", vol])

                QMessageBox.information(self, "Export", f"Exported to {path}")

            elif fmt == "json":
                path = export_dir / f"gymos_export_{date.today()}.json"
                from src.models import WeightEntry, NutritionLog, WaterLog, WorkoutSession, WorkoutExercise, ExerciseSet, Exercise

                weights = [{"date": str(w.date), "weight_kg": w.weight_kg} for w in session.query(WeightEntry).filter(WeightEntry.user_id == user.id).all()]
                nutrition = [{"date": str(n.date), "meal": n.meal, "calories": n.calories, "protein": n.protein_g, "fat": n.fat_g, "carbs": n.carbs_g} for n in session.query(NutritionLog).filter(NutritionLog.user_id == user.id).all()]

                data = {"user": {"name": user.name, "age": user.age, "height": user.height_cm, "goal": user.goal}, "weights": weights, "nutrition": nutrition}
                with open(path, "w") as f:
                    json.dump(data, f, indent=2)
                QMessageBox.information(self, "Export", f"Exported to {path}")

            elif fmt == "excel":
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment

                    path = export_dir / f"gymos_export_{date.today()}.xlsx"
                    wb = openpyxl.Workbook()

                    ws = wb.active
                    ws.title = "Weight"
                    ws.append(["Date", "Weight (kg)"])
                    for w in session.query(WeightEntry).filter(WeightEntry.user_id == user.id).order_by(WeightEntry.date).all():
                        ws.append([str(w.date), w.weight_kg])

                    ws2 = wb.create_sheet("Nutrition")
                    ws2.append(["Date", "Meal", "Calories", "Protein", "Fat", "Carbs"])
                    for n in session.query(NutritionLog).filter(NutritionLog.user_id == user.id).order_by(NutritionLog.date).all():
                        ws2.append([str(n.date), n.meal, n.calories, n.protein_g, n.fat_g, n.carbs_g])

                    wb.save(path)
                    QMessageBox.information(self, "Export", f"Exported to {path}")
                except ImportError:
                    QMessageBox.warning(self, "Export", "openpyxl not installed. Install with: pip install openpyxl")
        finally:
            session.close()

    def on_show(self):
        user = UserService.get_or_create_user()
        self.name_input.setText(user.name)
        self.age_input.setValue(user.age)
        self.height_input.setValue(user.height_cm)
        self.goal_input.setCurrentText(user.goal)
        self.split_input.setCurrentText(user.workout_split)
        self.target_weight_input.setValue(user.target_weight_kg)
        self.target_cal_input.setValue(user.target_calories)
        self.target_protein_input.setValue(user.target_protein_g)
        self.target_water_input.setValue(user.target_water_ml)
        self.days_input.setValue(user.workout_days_per_week)
