"""ProgramImporter — parse Excel, YAML, or JSON files into WorkoutProgram domain models."""

import json
import re
from pathlib import Path

from modules.workout_program.domain import (
    DeloadWeek,
    ProgramDay,
    ProgramExercise,
    ProgressionStrategy,
    WorkoutProgram,
)


class ProgramImporter:
    SUPPORTED_EXTENSIONS = {".xlsx", ".yaml", ".yml", ".json"}

    def import_file(self, filepath: str) -> WorkoutProgram:
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        ext = path.suffix.lower()
        if ext == ".xlsx":
            return self._import_excel(path)
        elif ext in (".yaml", ".yml"):
            return self._import_yaml(path)
        elif ext == ".json":
            return self._import_json(path)
        else:
            raise ValueError(
                f"Unsupported file format: {ext}. "
                f"Supported: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )

    def _program_name_from_path(self, path: Path) -> str:
        stem = path.stem
        stem = re.sub(r"[\s_-]+", " ", stem).strip()
        stem = stem.title()
        return stem

    # ─── Excel (.xlsx) ──────────────────────────────────────

    def _import_excel(self, path: Path) -> WorkoutProgram:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to import .xlsx files. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(str(path), data_only=True)

        sheet_name = self._find_tracker_sheet(wb)
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

        days: list[ProgramDay] = []
        current_day: ProgramDay | None = None

        for row in rows:
            vals = [str(v).strip() if v is not None else "" for v in row] + [""] * 8
            day_raw, exercise_raw, sets_raw, reps_raw = vals[:4]

            if not exercise_raw:
                continue

            if day_raw and (current_day is None or day_raw != current_day.name):
                current_day = ProgramDay(
                    name=day_raw,
                    sort_order=len(days),
                )
                days.append(current_day)

            if current_day is None:
                continue

            try:
                target_sets = int(sets_raw) if sets_raw else 3
            except (ValueError, TypeError):
                target_sets = 3

            target_reps = reps_raw if reps_raw else "10"

            current_day.exercises.append(ProgramExercise(
                name=exercise_raw,
                target_sets=target_sets,
                target_reps=target_reps,
                sort_order=len(current_day.exercises),
            ))

        rules: list[str] = []
        if "Rules" in wb.sheetnames:
            rules_ws = wb["Rules"]
            for row in rules_ws.iter_rows(min_row=1, max_row=rules_ws.max_row, values_only=True):
                val = str(row[0]).strip() if row[0] else ""
                if val:
                    rules.append(val)

        wb.close()

        return WorkoutProgram(
            name=self._program_name_from_path(path),
            description=f"Imported from {path.name}",
            source_file=path.name,
            rules=rules,
            days=days,
        )

    def _find_tracker_sheet(self, wb) -> str:
        candidates = ["Routine + Tracker", "Routine", "Tracker", "Workout", "Program"]
        for name in candidates:
            if name in wb.sheetnames:
                return name
        return wb.sheetnames[0]

    # ─── YAML (.yaml, .yml) ─────────────────────────────────

    def _import_yaml(self, path: Path) -> WorkoutProgram:
        try:
            import yaml
        except ImportError:
            raise ImportError("PyYAML is required to import .yaml files. Install with: pip install pyyaml")

        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f)

        if not isinstance(data, dict):
            raise ValueError("YAML file must contain a mapping at the top level.")

        return self._parse_dict(data, path.name)

    # ─── JSON (.json) ────────────────────────────────────────

    def _import_json(self, path: Path) -> WorkoutProgram:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            raise ValueError("JSON file must contain an object at the top level.")

        return self._parse_dict(data, path.name)

    def _parse_dict(self, data: dict, source_name: str) -> WorkoutProgram:
        days_data = data.get("days", [])
        days: list[ProgramDay] = []

        for d_idx, day_item in enumerate(days_data):
            if isinstance(day_item, str):
                day_item = {"name": day_item, "exercises": []}

            ex_list = day_item.get("exercises", [])
            exercises: list[ProgramExercise] = []
            for e_idx, ex_item in enumerate(ex_list):
                if isinstance(ex_item, str):
                    ex_item = {"name": ex_item}

                exercises.append(ProgramExercise(
                    name=ex_item["name"],
                    target_sets=ex_item.get("target_sets", 3),
                    target_reps=str(ex_item.get("target_reps", "10")),
                    muscle_group=ex_item.get("muscle_group", ""),
                    exercise_id=ex_item.get("exercise_id", ""),
                    sort_order=e_idx,
                    notes=ex_item.get("notes", ""),
                ))

            days.append(ProgramDay(
                name=day_item["name"],
                sort_order=d_idx,
                exercises=exercises,
                notes=day_item.get("notes", ""),
            ))

        deload_data = data.get("deload_week")
        deload_week = DeloadWeek(**deload_data) if deload_data and isinstance(deload_data, dict) else None

        prog_data = data.get("progression_strategy")
        progression_strategy = ProgressionStrategy(**prog_data) if prog_data and isinstance(prog_data, dict) else None

        return WorkoutProgram(
            name=data.get("name", self._program_name_from_path(Path(source_name))),
            description=data.get("description", f"Imported from {source_name}"),
            version=data.get("version", ""),
            author=data.get("author", ""),
            source_file=source_name,
            goal=data.get("goal", "hypertrophy"),
            experience_level=data.get("experience_level", "intermediate"),
            split=data.get("split", ""),
            mesocycle_duration_weeks=data.get("mesocycle_duration_weeks", 8),
            deload_week=deload_week,
            progression_strategy=progression_strategy,
            priority_muscles=data.get("priority_muscles", []),
            rules=data.get("rules", []),
            days=days,
        )

    # ─── Preview ─────────────────────────────────────────────

    def preview(self, filepath: str) -> dict:
        program = self.import_file(filepath)
        return {
            "name": program.name,
            "description": program.description,
            "source_file": program.source_file,
            "day_count": len(program.days),
            "exercise_count": sum(len(d.exercises) for d in program.days),
            "days": [
                {
                    "name": d.name,
                    "exercise_count": len(d.exercises),
                    "exercises": [e.name for e in d.exercises],
                }
                for d in program.days
            ],
        }
