"""Parse PPL_UL_MASTER_v6_UPDATED_LEGS.xlsx into canonical program.json.

Usage:
    python scripts/parse_excel_program.py [--excel PATH] [--output PATH]

This script is the ONLY way to regenerate the canonical program file.
The JSON output is the Source of Truth for GymOS workout programs.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# Muscle group mapping for all exercises in the canonical program.
# This must be kept in sync with the Excel file.
EXERCISE_MUSCLE_MAP: dict[str, str] = {
    # PUSH
    "Machine Chest Press": "chest",
    "Incline DB Press": "chest",
    "Pec Deck": "chest",
    "Machine Shoulder Press": "shoulders",
    "Lateral Raise": "shoulders",
    "Cable Tricep Extension": "triceps",
    # PULL
    "Lat Pulldown": "back",
    "Chest Supported Row": "back",
    "Seated Cable Row": "back",
    "Rear Delt Fly": "shoulders",
    "DB Curl": "biceps",
    "Rope Hammer Curl": "biceps",
    # LEGS
    "Leg Press": "quads",
    "Lying Leg Curl": "hamstrings",
    "Leg Extension": "quads",
    "Hip Abduction Machine": "glutes",
    "Standing Calf Raise": "calves",
    "Cable Crunch": "abs",
    # UPPER
    "Incline Machine Press": "chest",
    "Machine Row": "back",
    "Cable Curl": "biceps",
    "Tricep Extension": "triceps",
    # LOWER
    "Hack Squat": "quads",
    "Hanging Knee Raise": "abs",
}


def parse_program(filepath: str) -> dict:
    """Parse the Excel file and return a program dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if "Routine + Tracker" not in wb.sheetnames:
        print(f"Error: 'Routine + Tracker' sheet not found in {filepath}")
        sys.exit(1)

    ws = wb["Routine + Tracker"]

    # Read all rows, skip header
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

    days: list[dict] = []
    current_day: dict | None = None
    current_day_name: str = ""

    for row in rows:
        day_raw, exercise_raw, sets_raw, reps_raw, *_ = (
            [str(v).strip() if v is not None else "" for v in row] + [""] * 8
        )

        if not exercise_raw:
            continue

        # A new day starts only when the Day column changes to a different value
        if day_raw and day_raw != current_day_name:
            current_day_name = day_raw
            current_day = {
                "name": day_raw,
                "exercises": [],
            }
            days.append(current_day)

        if current_day is None:
            continue

        # Parse sets
        try:
            target_sets = int(sets_raw) if sets_raw else 3
        except ValueError:
            target_sets = 3

        # Parse reps — could be "8-10", "10", "12-15", etc.
        target_reps = reps_raw if reps_raw else "10"

        # Determine muscle group
        muscle_group = EXERCISE_MUSCLE_MAP.get(exercise_raw, "")

        current_day["exercises"].append({
            "name": exercise_raw,
            "target_sets": target_sets,
            "target_reps": target_reps,
            "muscle_group": muscle_group,
        })

    # Get the rules from the Rules sheet
    rules: list[str] = []
    if "Rules" in wb.sheetnames:
        rules_ws = wb["Rules"]
        for row in rules_ws.iter_rows(min_row=1, max_row=rules_ws.max_row, values_only=True):
            val = str(row[0]).strip() if row[0] else ""
            if val:
                rules.append(val)

    wb.close()

    return {
        "name": "PPL-UL v6",
        "description": "PPL-UL Master v6 — Updated Legs",
        "source": Path(filepath).name,
        "rules": rules,
        "days": days,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse PPL-UL Excel into program JSON")
    parser.add_argument(
        "--excel",
        default="D:/Personal/PPL_UL_MASTER_v6_UPDATED_LEGS.xlsx",
        help="Path to the canonical Excel file",
    )
    parser.add_argument(
        "--output",
        default="data/program.json",
        help="Output JSON file path",
    )
    args = parser.parse_args()

    program = parse_program(args.excel)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(program, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    day_count = len(program["days"])
    exercise_count = sum(len(d["exercises"]) for d in program["days"])
    print(f"OK Generated {output_path}")
    print(f"   Days: {day_count}")
    print(f"   Exercises: {exercise_count}")


if __name__ == "__main__":
    main()
